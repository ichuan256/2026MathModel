from __future__ import annotations

import csv
import math
from pathlib import Path

import numpy as np
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import matplotlib.pyplot as plt

    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

try:
    from PIL import ImageFont

    HAS_PIL = True
except ImportError:
    HAS_PIL = False


BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "附件.xlsx"
POINTS_CSV_PATH = BASE_DIR / "attachment_bathymetry_points.csv"
CONTOUR_PNG_PATH = BASE_DIR / "attachment_contour_map.png"
CONTOUR_SVG_PATH = BASE_DIR / "attachment_contour_map.svg"


def is_number(value) -> bool:
    return isinstance(value, (int, float)) and math.isfinite(float(value))


def find_x_axis(rows: list[list]) -> tuple[int, list[int], list[float]]:
    best = None
    for row_index, row in enumerate(rows[:20]):
        numeric_cols = [(col_index, float(value)) for col_index, value in enumerate(row) if is_number(value)]
        if len(numeric_cols) < 10:
            continue
        values = [value for _, value in numeric_cols]
        if max(values) - min(values) > 20:
            continue
        increasing_pairs = sum(b > a for a, b in zip(values, values[1:]))
        if increasing_pairs >= len(values) * 0.8:
            diffs = [b - a for a, b in zip(values, values[1:]) if b > a]
            mean_step = sum(diffs) / max(len(diffs), 1)
            step_error = sum(abs(d - mean_step) for d in diffs) / max(len(diffs), 1)
            score = (len(numeric_cols), -step_error)
            if best is None or score > best[0]:
                best = (score, row_index, [col for col, _ in numeric_cols], values)
    if best is None:
        raise ValueError("无法识别东西方向坐标行。")
    _, row_index, x_cols, values = best
    return row_index, x_cols, values


def find_y_value(row: list, first_x_col: int) -> float | None:
    for value in row[:first_x_col]:
        if is_number(value):
            return float(value)
    return None


def read_bathymetry_grid() -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    print("[1/4] 读取附件并识别测深网格...", flush=True)
    if openpyxl is None:
        if not POINTS_CSV_PATH.exists():
            raise FileNotFoundError(f"Neither openpyxl nor cached point data is available: {POINTS_CSV_PATH}")
        rows = []
        with POINTS_CSV_PATH.open("r", encoding="utf-8-sig", newline="") as file:
            for row in csv.DictReader(file):
                rows.append((float(row["x_nm"]), float(row["y_nm"]), float(row["depth_m"])))
        x = np.unique(np.asarray([row[0] for row in rows], dtype=float))
        y = np.unique(np.asarray([row[1] for row in rows], dtype=float))
        depth = np.full((len(y), len(x)), np.nan, dtype=float)
        x_index = {value: index for index, value in enumerate(x)}
        y_index = {value: index for index, value in enumerate(y)}
        for x_value, y_value, z_value in rows:
            depth[y_index[y_value], x_index[x_value]] = z_value
        print(f"    使用缓存 CSV；x节点={len(x)}, y节点={len(y)}, 有效测深点={np.isfinite(depth).sum()}", flush=True)
        return x, y, depth
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]

    x_row_index, x_cols, x_values = find_x_axis(rows)
    first_x_col = min(x_cols)
    y_values: list[float] = []
    depth_rows: list[list[float]] = []

    for row in rows[x_row_index + 1 :]:
        y = find_y_value(row, first_x_col)
        if y is None:
            continue
        depths = []
        valid_count = 0
        for col in x_cols:
            value = row[col] if col < len(row) else None
            if is_number(value):
                depths.append(float(value))
                valid_count += 1
            else:
                depths.append(np.nan)
        if valid_count >= max(5, len(x_cols) * 0.5):
            y_values.append(y)
            depth_rows.append(depths)

    if not y_values or not depth_rows:
        raise ValueError("无法识别南北坐标和测深矩阵。")

    x = np.asarray(x_values, dtype=float)
    y = np.asarray(y_values, dtype=float)
    depth = np.asarray(depth_rows, dtype=float)
    print(f"    x节点={len(x)}, y节点={len(y)}, 有效测深点={np.isfinite(depth).sum()}", flush=True)
    return x, y, depth


def export_points(x: np.ndarray, y: np.ndarray, depth: np.ndarray) -> None:
    print("[2/4] 导出离散测深点 CSV...", flush=True)
    with POINTS_CSV_PATH.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["x_nm", "y_nm", "depth_m"])
        for row_index, y_value in enumerate(y):
            for col_index, x_value in enumerate(x):
                z = depth[row_index, col_index]
                if np.isfinite(z):
                    writer.writerow([round(float(x_value), 6), round(float(y_value), 6), round(float(z), 4)])


def nice_levels(depth: np.ndarray, interval: float = 10.0) -> np.ndarray:
    z_min = float(np.nanmin(depth))
    z_max = float(np.nanmax(depth))
    start = math.floor(z_min / interval) * interval
    stop = math.ceil(z_max / interval) * interval
    return np.arange(start, stop + interval, interval)


def load_font(size: int, bold: bool = False):
    if not HAS_PIL:
        return None
    candidates = [
        Path("C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf"),
    ]
    for font_path in candidates:
        if font_path.exists():
            return ImageFont.truetype(str(font_path), size=size)
    return ImageFont.load_default()


def marching_segments(
    x: np.ndarray, y: np.ndarray, depth: np.ndarray, level: float
) -> list[tuple[tuple[float, float], tuple[float, float]]]:
    segments = []

    def interp(p1, z1, p2, z2):
        if abs(z2 - z1) < 1e-12:
            return p1
        t = (level - z1) / (z2 - z1)
        return (p1[0] + t * (p2[0] - p1[0]), p1[1] + t * (p2[1] - p1[1]))

    for i in range(len(y) - 1):
        for j in range(len(x) - 1):
            corners = [
                ((x[j], y[i]), depth[i, j]),
                ((x[j + 1], y[i]), depth[i, j + 1]),
                ((x[j + 1], y[i + 1]), depth[i + 1, j + 1]),
                ((x[j], y[i + 1]), depth[i + 1, j]),
            ]
            if any(not np.isfinite(z) for _, z in corners):
                continue
            crossings = []
            for a, b in ((0, 1), (1, 2), (2, 3), (3, 0)):
                p1, z1 = corners[a]
                p2, z2 = corners[b]
                if (z1 - level) * (z2 - level) < 0:
                    crossings.append(interp(p1, z1, p2, z2))
            if len(crossings) == 2:
                segments.append((crossings[0], crossings[1]))
            elif len(crossings) == 4:
                segments.append((crossings[0], crossings[1]))
                segments.append((crossings[2], crossings[3]))
    return segments


def setup_matplotlib() -> None:
    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["svg.fonttype"] = "none"
    plt.rcParams["path.simplify"] = False


def plot_contour(x: np.ndarray, y: np.ndarray, depth: np.ndarray) -> None:
    if not HAS_MATPLOTLIB:
        raise ImportError("需要安装 matplotlib 才能输出平滑 PNG/SVG。")

    print("[3/4] 绘制横向拉伸版等深线 PNG/SVG...", flush=True)
    setup_matplotlib()
    x_grid, y_grid = np.meshgrid(x, y)
    levels = nice_levels(depth)
    z_min, z_max = float(np.nanmin(depth)), float(np.nanmax(depth))

    fig = plt.figure(figsize=(12.6, 7.0), facecolor="white")
    ax = fig.add_axes([0.075, 0.12, 0.82, 0.82])

    filled = ax.contourf(x_grid, y_grid, depth, levels=levels, cmap="YlGnBu", extend="both")
    contour = ax.contour(x_grid, y_grid, depth, levels=levels, colors="#18384a", linewidths=0.42, alpha=0.84)
    ax.clabel(contour, inline=False, fmt="%.0f m", fontsize=8)
    ax.scatter(x_grid.ravel(), y_grid.ravel(), s=1.6, c="white", alpha=0.11, linewidths=0)

    colorbar = fig.colorbar(filled, ax=ax, pad=0.014, shrink=0.94)
    colorbar.set_label("海水深度 / m")
    ax.set_xlabel("东西方向 / n mile", fontsize=13, fontweight="bold")
    ax.set_ylabel("南北方向 / n mile", fontsize=13, fontweight="bold")
    ax.set_aspect("auto")
    ax.grid(color="white", linewidth=0.55, alpha=0.42)

    fig.savefig(CONTOUR_PNG_PATH, dpi=300, bbox_inches="tight", facecolor="white")
    fig.savefig(CONTOUR_SVG_PATH, format="svg", bbox_inches="tight", facecolor="white")
    plt.close(fig)


def main() -> None:
    print("=== 2023B 问题四：离散测深点等深线图 ===", flush=True)
    x, y, depth = read_bathymetry_grid()
    export_points(x, y, depth)
    plot_contour(x, y, depth)
    print("[4/4] 完成。", flush=True)
    print("saved_points:", POINTS_CSV_PATH)
    print("saved_png:", CONTOUR_PNG_PATH)
    print("saved_svg:", CONTOUR_SVG_PATH)


if __name__ == "__main__":
    main()
