from __future__ import annotations

import html
import math
from pathlib import Path

from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont


BASE_DIR = Path(__file__).resolve().parent
PROBLEM_DIR = BASE_DIR

OUTPUT_DIR = PROBLEM_DIR
XLSX_PATH = OUTPUT_DIR / "q2_coverage_result.xlsx"
PNG_PATH = OUTPUT_DIR / "q2_coverage_heatmap.png"
SVG_PATH = OUTPUT_DIR / "q2_coverage_heatmap.svg"
RESULT2_TEMPLATE_PATH = PROBLEM_DIR / "result2.xlsx"

OPEN_ANGLE_DEG = 120.0
SLOPE_DEG = 1.5
CENTER_DEPTH_M = 120.0
NAUTICAL_MILE_M = 1852.0

DISTANCES_NM = [0, 0.3, 0.6, 0.9, 1.2, 1.5, 1.8, 2.1]
DIRECTION_DEG = [0, 45, 90, 135, 180, 225, 270, 315]


def q2_width(distance_nm: float, beta_deg: float) -> dict[str, float]:
    theta_half = math.radians(OPEN_ANGLE_DEG / 2.0)
    alpha = math.radians(SLOPE_DEG)
    beta = math.radians(beta_deg)
    distance_m = distance_nm * NAUTICAL_MILE_M

    depth = CENTER_DEPTH_M - distance_m * math.cos(beta) * math.tan(alpha)
    alpha_beta = math.atan(math.sin(beta) * math.tan(alpha))

    left = depth * math.sin(theta_half) / math.cos(theta_half + alpha_beta)
    right = depth * math.sin(theta_half) / math.cos(theta_half - alpha_beta)
    slope_width = left + right
    horizontal_width = slope_width * math.cos(alpha_beta)

    return {
        "depth_m": depth,
        "alpha_beta_deg": math.degrees(alpha_beta),
        "slope_width_m": slope_width,
        "horizontal_width_m": horizontal_width,
    }


def compute_table() -> list[list[dict[str, float]]]:
    return [[q2_width(distance, beta) for distance in DISTANCES_NM] for beta in DIRECTION_DEG]


def write_xlsx(table: list[list[dict[str, float]]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "问题二覆盖宽度"
    ws.cell(row=1, column=1, value="覆盖宽度/m")
    ws.cell(row=1, column=3, value="测量船距海域中心点处的距离/海里")
    for col, distance in enumerate(DISTANCES_NM, start=3):
        ws.cell(row=2, column=col, value=distance)
    ws.cell(row=3, column=1, value="测线方向夹角/度")
    for row, beta in enumerate(DIRECTION_DEG, start=3):
        ws.cell(row=row, column=2, value=beta)
        for col, item in enumerate(table[row - 3], start=3):
            ws.cell(row=row, column=col, value=round(item["slope_width_m"], 2))

    ws2 = wb.create_sheet("水平投影宽度")
    ws2.cell(row=1, column=1, value="水平投影覆盖宽度/m")
    ws2.cell(row=1, column=3, value="测量船距海域中心点处的距离/海里")
    for col, distance in enumerate(DISTANCES_NM, start=3):
        ws2.cell(row=2, column=col, value=distance)
    ws2.cell(row=3, column=1, value="测线方向夹角/度")
    for row, beta in enumerate(DIRECTION_DEG, start=3):
        ws2.cell(row=row, column=2, value=beta)
        for col, item in enumerate(table[row - 3], start=3):
            ws2.cell(row=row, column=col, value=round(item["horizontal_width_m"], 2))

    wb.save(XLSX_PATH)

    if RESULT2_TEMPLATE_PATH.exists():
        from openpyxl import load_workbook

        template = load_workbook(RESULT2_TEMPLATE_PATH)
        sheet = template.active
        for row, beta in enumerate(DIRECTION_DEG, start=3):
            sheet.cell(row=row, column=2, value=beta)
            for col, item in enumerate(table[row - 3], start=3):
                sheet.cell(row=row, column=col, value=round(item["slope_width_m"], 2))
        template.save(RESULT2_TEMPLATE_PATH)


def load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        Path("C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/simsun.ttc"),
    ]
    for font_path in candidates:
        if font_path.exists():
            return ImageFont.truetype(str(font_path), size=size)
    return ImageFont.load_default()


def blend(a: tuple[int, int, int], b: tuple[int, int, int], t: float) -> tuple[int, int, int]:
    return tuple(round(a[i] + (b[i] - a[i]) * t) for i in range(3))


def heat_color(value: float, min_value: float, max_value: float) -> tuple[int, int, int]:
    t = (value - min_value) / (max_value - min_value)
    if t < 0.5:
        return blend((224, 242, 241), (77, 182, 172), t * 2)
    return blend((77, 182, 172), (0, 105, 92), (t - 0.5) * 2)


def draw_centered(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    text: str,
    font: ImageFont.ImageFont,
    fill: tuple[int, int, int],
) -> None:
    bbox = draw.textbbox((0, 0), text, font=font)
    x = box[0] + (box[2] - box[0] - (bbox[2] - bbox[0])) / 2
    y = box[1] + (box[3] - box[1] - (bbox[3] - bbox[1])) / 2
    draw.text((x, y), text, font=font, fill=fill)


def write_png(table: list[list[dict[str, float]]]) -> None:
    values = [[item["slope_width_m"] for item in row] for row in table]
    flat = [value for row in values for value in row]
    min_value, max_value = min(flat), max(flat)

    left_w, top_h, cell_w, cell_h = 150, 162, 112, 62
    right_pad, bottom_pad = 44, 120
    width = left_w + cell_w * len(DISTANCES_NM) + right_pad
    height = top_h + cell_h * len(DIRECTION_DEG) + bottom_pad

    image = Image.new("RGB", (width, height), (248, 250, 252))
    draw = ImageDraw.Draw(image)
    title_font = load_font(30, bold=True)
    label_font = load_font(18, bold=True)
    cell_font = load_font(17)
    small_font = load_font(14)

    draw.text((36, 26), "问题二多波束覆盖宽度计算结果", font=title_font, fill=(20, 32, 44))
    draw.text(
        (36, 78),
        "主表采用斜面实际覆盖宽度 Ws；参数：开角120°, 坡度1.5°, 中心水深120 m",
        font=small_font,
        fill=(71, 85, 105),
    )

    draw.rectangle((left_w, top_h - cell_h, left_w + cell_w * len(DISTANCES_NM), top_h), fill=(226, 232, 240))
    draw.rectangle((0, top_h, left_w, top_h + cell_h * len(DIRECTION_DEG)), fill=(226, 232, 240))
    draw_centered(draw, (0, top_h - cell_h, left_w, top_h), "夹角β/°", label_font, (30, 41, 59))
    for col, distance in enumerate(DISTANCES_NM):
        x0 = left_w + col * cell_w
        draw_centered(draw, (x0, top_h - cell_h, x0 + cell_w, top_h), f"{distance:g} 海里", label_font, (30, 41, 59))

    for row, beta in enumerate(DIRECTION_DEG):
        y0 = top_h + row * cell_h
        draw_centered(draw, (0, y0, left_w, y0 + cell_h), str(beta), label_font, (30, 41, 59))
        for col, value in enumerate(values[row]):
            x0 = left_w + col * cell_w
            fill = heat_color(value, min_value, max_value)
            draw.rectangle((x0, y0, x0 + cell_w, y0 + cell_h), fill=fill)
            text_fill = (248, 250, 252) if value > (min_value + max_value) / 2 else (15, 23, 42)
            draw_centered(draw, (x0, y0, x0 + cell_w, y0 + cell_h), f"{value:.2f}", cell_font, text_fill)

    grid_color = (148, 163, 184)
    for col in range(len(DISTANCES_NM) + 1):
        x = left_w + col * cell_w
        draw.line((x, top_h - cell_h, x, top_h + cell_h * len(DIRECTION_DEG)), fill=grid_color, width=1)
    for row in range(len(DIRECTION_DEG) + 2):
        y = top_h - cell_h + row * cell_h
        draw.line((0, y, left_w + cell_w * len(DISTANCES_NM), y), fill=grid_color, width=1)
    draw.line((left_w, top_h - cell_h, left_w + cell_w * len(DISTANCES_NM), top_h - cell_h), fill=grid_color, width=1)
    draw.line((0, top_h - cell_h, 0, top_h + cell_h * len(DIRECTION_DEG)), fill=grid_color, width=1)

    legend_x, legend_y = 36, height - 74
    draw.text((legend_x, legend_y - 24), "颜色越深表示覆盖宽度越大，单位：m", font=small_font, fill=(71, 85, 105))
    for i in range(220):
        t_value = min_value + (max_value - min_value) * i / 219
        draw.line((legend_x + i, legend_y, legend_x + i, legend_y + 16), fill=heat_color(t_value, min_value, max_value), width=1)
    draw.text((legend_x, legend_y + 22), f"{min_value:.2f}", font=small_font, fill=(71, 85, 105))
    draw.text((legend_x + 168, legend_y + 22), f"{max_value:.2f}", font=small_font, fill=(71, 85, 105))

    image.save(PNG_PATH)


def write_svg(table: list[list[dict[str, float]]]) -> None:
    values = [[item["slope_width_m"] for item in row] for row in table]
    flat = [value for row in values for value in row]
    min_value, max_value = min(flat), max(flat)

    left_w, top_h, cell_w, cell_h = 150, 162, 112, 62
    right_pad, bottom_pad = 44, 120
    width = left_w + cell_w * len(DISTANCES_NM) + right_pad
    height = top_h + cell_h * len(DIRECTION_DEG) + bottom_pad

    def color(value: float) -> str:
        r, g, b = heat_color(value, min_value, max_value)
        return f"rgb({r},{g},{b})"

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        "<style><![CDATA[text{font-family:'Microsoft YaHei','SimHei',Arial,sans-serif}.thin{vector-effect:non-scaling-stroke}.fixed-label{font-size:14px}.title{font-size:30px;font-weight:700}.head{font-size:18px;font-weight:700}.cell{font-size:17px}]]></style>",
        f'<rect width="{width}" height="{height}" fill="#f8fafc"/>',
        '<text x="36" y="58" class="title" fill="#14202c">问题二：多波束覆盖宽度计算结果</text>',
        '<text x="36" y="92" class="fixed-label" fill="#475569">主表采用斜面实际覆盖宽度 Ws；参数：开角120°，坡度1.5°，中心水深120 m</text>',
        f'<rect x="{left_w}" y="{top_h - cell_h}" width="{cell_w * len(DISTANCES_NM)}" height="{cell_h}" fill="#e2e8f0"/>',
        f'<rect x="0" y="{top_h}" width="{left_w}" height="{cell_h * len(DIRECTION_DEG)}" fill="#e2e8f0"/>',
    ]

    def text_center(box: tuple[float, float, float, float], content: str, cls: str, fill: str = "#1e293b") -> None:
        x = (box[0] + box[2]) / 2
        y = (box[1] + box[3]) / 2 + 6
        parts.append(f'<text x="{x:.2f}" y="{y:.2f}" class="{cls}" fill="{fill}" text-anchor="middle">{html.escape(content)}</text>')

    text_center((0, top_h - cell_h, left_w, top_h), "夹角β/°", "head")
    for col, distance in enumerate(DISTANCES_NM):
        x0 = left_w + col * cell_w
        text_center((x0, top_h - cell_h, x0 + cell_w, top_h), f"{distance:g} 海里", "head")

    for row, beta in enumerate(DIRECTION_DEG):
        y0 = top_h + row * cell_h
        text_center((0, y0, left_w, y0 + cell_h), str(beta), "head")
        for col, value in enumerate(values[row]):
            x0 = left_w + col * cell_w
            fill = color(value)
            text_fill = "#f8fafc" if value > (min_value + max_value) / 2 else "#0f172a"
            parts.append(f'<rect x="{x0}" y="{y0}" width="{cell_w}" height="{cell_h}" fill="{fill}"/>')
            text_center((x0, y0, x0 + cell_w, y0 + cell_h), f"{value:.2f}", "cell", text_fill)

    grid_color = "#94a3b8"
    for col in range(len(DISTANCES_NM) + 1):
        x = left_w + col * cell_w
        parts.append(f'<line class="thin" x1="{x}" y1="{top_h - cell_h}" x2="{x}" y2="{top_h + cell_h * len(DIRECTION_DEG)}" stroke="{grid_color}" stroke-width="1"/>')
    for row in range(len(DIRECTION_DEG) + 2):
        y = top_h - cell_h + row * cell_h
        parts.append(f'<line class="thin" x1="0" y1="{y}" x2="{left_w + cell_w * len(DISTANCES_NM)}" y2="{y}" stroke="{grid_color}" stroke-width="1"/>')

    legend_x, legend_y = 36, height - 74
    parts.append(f'<text x="{legend_x}" y="{legend_y - 24}" class="fixed-label" fill="#475569">颜色越深表示覆盖宽度越大，单位：m</text>')
    for i in range(220):
        t_value = min_value + (max_value - min_value) * i / 219
        parts.append(f'<line class="thin" x1="{legend_x + i}" y1="{legend_y}" x2="{legend_x + i}" y2="{legend_y + 16}" stroke="{color(t_value)}" stroke-width="1"/>')
    parts.append(f'<text x="{legend_x}" y="{legend_y + 38}" class="fixed-label" fill="#475569">{min_value:.2f}</text>')
    parts.append(f'<text x="{legend_x + 168}" y="{legend_y + 38}" class="fixed-label" fill="#475569">{max_value:.2f}</text>')
    parts.append("</svg>")
    SVG_PATH.write_text("\n".join(parts), encoding="utf-8")


def main() -> None:
    table = compute_table()
    write_xlsx(table)
    write_png(table)
    write_svg(table)
    print(f"xlsx: {XLSX_PATH}")
    if RESULT2_TEMPLATE_PATH.exists():
        print(f"template: {RESULT2_TEMPLATE_PATH}")
    print(f"png: {PNG_PATH}")
    print(f"svg: {SVG_PATH}")
    print("Ws table:")
    print("beta\\distance", *DISTANCES_NM)
    for beta, row in zip(DIRECTION_DEG, table):
        print(beta, *[f"{item['slope_width_m']:.2f}" for item in row])


if __name__ == "__main__":
    main()
