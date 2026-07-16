from math import cos, pi, sin, tan
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


THETA = 120 * pi / 180
GAMMA = THETA / 2
ALPHA = 1.5 * pi / 180
CENTER_DEPTH = 70.0
LINE_SPACING = 200.0
DISTANCES = [-800, -600, -400, -200, 0, 200, 400, 600, 800]

BASE_DIR = Path(__file__).resolve().parent
PROBLEM_DIR = BASE_DIR
WORKBOOK_PATH = PROBLEM_DIR / "result1.xlsx"
PNG_PATH = PROBLEM_DIR / "q1_multibeam_profile.png"
SVG_PATH = PROBLEM_DIR / "q1_multibeam_profile.svg"


def depth_at(distance: float) -> float:
    return CENTER_DEPTH - distance * tan(ALPHA)


def side_widths(depth: float) -> tuple[float, float]:
    common = depth * sin(GAMMA) * cos(ALPHA)
    left_width = common / cos(GAMMA + ALPHA)
    right_width = common / cos(GAMMA - ALPHA)
    return left_width, right_width


def total_width(depth: float) -> float:
    left_width, right_width = side_widths(depth)
    return left_width + right_width


def round2(value: float) -> float:
    return round(value + 1e-12, 2)


def load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        Path("C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/simsun.ttc"),
    ]
    for font_path in candidates:
        if font_path.exists():
            return ImageFont.truetype(str(font_path), size=size)
    return ImageFont.load_default()


def solve() -> tuple[list[float], list[float], list[float | str]]:
    depths = [depth_at(distance) for distance in DISTANCES]
    widths = [total_width(depth) for depth in depths]
    side_pairs = [side_widths(depth) for depth in depths]

    overlaps: list[float | str] = ["\u2014\u2014"]
    for index in range(1, len(DISTANCES)):
        previous_right = side_pairs[index - 1][1]
        current_left = side_pairs[index][0]
        overlap_length = previous_right + current_left - LINE_SPACING
        overlap_rate = overlap_length / widths[index] * 100
        overlaps.append(round2(overlap_rate))

    return [round2(x) for x in depths], [round2(x) for x in widths], overlaps


def draw_centered(
    draw: ImageDraw.ImageDraw,
    point: tuple[float, float],
    text: str,
    font: ImageFont.ImageFont,
    fill: tuple[int, int, int],
) -> None:
    bbox = draw.textbbox((0, 0), text, font=font)
    draw.text((point[0] - (bbox[2] - bbox[0]) / 2, point[1] - (bbox[3] - bbox[1]) / 2), text, font=font, fill=fill)


def map_point(distance: float, value: float, value_min: float, value_max: float) -> tuple[float, float]:
    left, top, plot_w, plot_h = 95, 72, 840, 438
    x = left + (distance - min(DISTANCES)) / (max(DISTANCES) - min(DISTANCES)) * plot_w
    y = top + (value_max - value) / max(value_max - value_min, 1e-12) * plot_h
    return x, y


def draw_png(depths: list[float], widths: list[float], overlaps: list[float | str]) -> None:
    image = Image.new("RGB", (1000, 620), (255, 255, 255))
    draw = ImageDraw.Draw(image, "RGBA")
    label_font = load_font(18, True)
    text_font = load_font(15)
    small_font = load_font(14)

    left, top, plot_w, plot_h = 95, 72, 840, 438
    right, bottom = left + plot_w, top + plot_h
    value_min, value_max = 40.0, 340.0
    depth_color = (49, 91, 145, 255)
    width_color = (200, 88, 35, 255)

    for value in [40, 100, 160, 220, 280, 340]:
        _, y = map_point(0, value, value_min, value_max)
        draw.line((left, y, right, y), fill=(210, 214, 220, 190), width=1)
        bbox = draw.textbbox((0, 0), str(value), font=small_font)
        draw.text((left - 14 - (bbox[2] - bbox[0]), y - 8), str(value), font=small_font, fill=(70, 75, 82))
    for distance in DISTANCES:
        x, _ = map_point(distance, value_min, value_min, value_max)
        draw_centered(draw, (x, bottom + 22), str(distance), small_font, (70, 75, 82))

    depth_points = [map_point(d, z, value_min, value_max) for d, z in zip(DISTANCES, depths)]
    width_points = [map_point(d, w, value_min, value_max) for d, w in zip(DISTANCES, widths)]
    draw.line(depth_points, fill=depth_color, width=4)
    draw.line(width_points, fill=width_color, width=4)
    for point, value in zip(depth_points, depths):
        draw.ellipse((point[0] - 5, point[1] - 5, point[0] + 5, point[1] + 5), fill=depth_color, outline=(255, 255, 255, 255), width=1)
        draw_centered(draw, (point[0], point[1] - 18), f"{value:.1f}", small_font, depth_color[:3])
    for point, value in zip(width_points, widths):
        draw.rectangle((point[0] - 5, point[1] - 5, point[0] + 5, point[1] + 5), fill=width_color, outline=(255, 255, 255, 255), width=1)
        draw_centered(draw, (point[0], point[1] + 20), f"{value:.1f}", small_font, (160, 62, 24))

    draw.line((left, top, left, bottom), fill=(35, 39, 45, 255), width=2)
    draw.line((left, bottom, right, bottom), fill=(35, 39, 45, 255), width=2)
    draw_centered(draw, (left + plot_w / 2, bottom + 64), "测线距中心距离 / m", label_font, (35, 39, 45))
    draw.text((left, 30), "水深与覆盖宽度 / m", font=label_font, fill=(35, 39, 45))

    legend_x, legend_y = right - 188, top + 16
    draw.rounded_rectangle((legend_x, legend_y, right - 16, legend_y + 76), radius=5, fill=(255, 255, 255, 235), outline=(185, 188, 193, 255), width=1)
    draw.line((legend_x + 16, legend_y + 24, legend_x + 54, legend_y + 24), fill=depth_color, width=4)
    draw.ellipse((legend_x + 30, legend_y + 19, legend_x + 40, legend_y + 29), fill=depth_color)
    draw.text((legend_x + 66, legend_y + 13), "海水深度", font=text_font, fill=(35, 39, 45))
    draw.line((legend_x + 16, legend_y + 54, legend_x + 54, legend_y + 54), fill=width_color, width=4)
    draw.rectangle((legend_x + 30, legend_y + 49, legend_x + 40, legend_y + 59), fill=width_color)
    draw.text((legend_x + 66, legend_y + 43), "覆盖宽度", font=text_font, fill=(35, 39, 45))

    image.save(PNG_PATH, dpi=(300, 300))


def draw_svg(depths: list[float], widths: list[float], overlaps: list[float | str]) -> None:
    left, top, plot_w, plot_h = 95, 72, 840, 438
    right, bottom = left + plot_w, top + plot_h
    value_min, value_max = 40.0, 340.0

    def pt(distance: float, value: float) -> tuple[float, float]:
        return map_point(distance, value, value_min, value_max)

    def points(values: list[float]) -> str:
        return " ".join(f"{pt(d, v)[0]:.2f},{pt(d, v)[1]:.2f}" for d, v in zip(DISTANCES, values))

    parts = [
        '<svg xmlns="http://www.w3.org/2000/svg" width="1000" height="620" viewBox="0 0 1000 620">',
        "<style><![CDATA[text{font-family:'Microsoft YaHei','SimHei',Arial,sans-serif}.thin{vector-effect:non-scaling-stroke}.fixed-label{font-size:14px}.axis{font-size:18px;font-weight:700}.panel{font-size:15px}]]></style>",
        '<rect width="1000" height="620" fill="#ffffff"/>',
    ]
    for i in range(6):
        y = top + i / 5 * plot_h
        value = value_max - i / 5 * (value_max - value_min)
        parts.append(f'<line class="thin" x1="{left}" y1="{y:.2f}" x2="{right}" y2="{y:.2f}" stroke="#cbd5e1" stroke-width="1" opacity="0.75"/>')
        parts.append(f'<text x="30" y="{y + 5:.2f}" class="fixed-label" fill="#475569">{value:.0f}</text>')
    for distance in DISTANCES:
        x, _ = pt(distance, value_min)
        parts.append(f'<line class="thin" x1="{x:.2f}" y1="{top}" x2="{x:.2f}" y2="{bottom}" stroke="#e2e8f0" stroke-width="1" opacity="0.7"/>')
        parts.append(f'<text x="{x:.2f}" y="{bottom + 22}" class="fixed-label" fill="#475569" text-anchor="middle">{distance}</text>')
    parts.append(f'<polyline class="thin" points="{points(depths)}" fill="none" stroke="#315b91" stroke-width="4" stroke-linecap="round" stroke-linejoin="round"/>')
    parts.append(f'<polyline class="thin" points="{points(widths)}" fill="none" stroke="#c85823" stroke-width="4" stroke-linecap="round" stroke-linejoin="round"/>')
    for distance, value in zip(DISTANCES, depths):
        x, y = pt(distance, value)
        parts.append(f'<circle cx="{x:.2f}" cy="{y:.2f}" r="5" fill="#315b91" stroke="#ffffff" stroke-width="1"/>')
        parts.append(f'<text x="{x:.2f}" y="{y - 14:.2f}" class="fixed-label" fill="#315b91" text-anchor="middle">{value:.1f}</text>')
    for distance, value in zip(DISTANCES, widths):
        x, y = pt(distance, value)
        parts.append(f'<rect x="{x - 5:.2f}" y="{y - 5:.2f}" width="10" height="10" fill="#c85823" stroke="#ffffff" stroke-width="1"/>')
        parts.append(f'<text x="{x:.2f}" y="{y + 24:.2f}" class="fixed-label" fill="#a03e18" text-anchor="middle">{value:.1f}</text>')
    parts.append(f'<line class="thin" x1="{left}" y1="{top}" x2="{left}" y2="{bottom}" stroke="#23272d" stroke-width="2"/>')
    parts.append(f'<line class="thin" x1="{left}" y1="{bottom}" x2="{right}" y2="{bottom}" stroke="#23272d" stroke-width="2"/>')
    parts.append(f'<text x="{left + plot_w / 2}" y="{bottom + 69}" class="axis" fill="#23272d" text-anchor="middle">测线距中心距离 / m</text>')
    parts.append(f'<text x="{left}" y="48" class="axis" fill="#23272d">水深与覆盖宽度 / m</text>')
    panel_x, panel_y = right - 188, top + 16
    parts.extend([
        f'<rect class="thin" x="{panel_x}" y="{panel_y}" width="172" height="76" rx="5" fill="#ffffff" fill-opacity="0.94" stroke="#b9bcc1"/>',
        f'<line class="thin" x1="{panel_x + 16}" y1="{panel_y + 24}" x2="{panel_x + 54}" y2="{panel_y + 24}" stroke="#315b91" stroke-width="4"/>',
        f'<circle cx="{panel_x + 35}" cy="{panel_y + 24}" r="5" fill="#315b91"/>',
        f'<text x="{panel_x + 66}" y="{panel_y + 29}" class="panel" fill="#23272d">海水深度</text>',
        f'<line class="thin" x1="{panel_x + 16}" y1="{panel_y + 54}" x2="{panel_x + 54}" y2="{panel_y + 54}" stroke="#c85823" stroke-width="4"/>',
        f'<rect x="{panel_x + 30}" y="{panel_y + 49}" width="10" height="10" fill="#c85823"/>',
        f'<text x="{panel_x + 66}" y="{panel_y + 59}" class="panel" fill="#23272d">覆盖宽度</text>',
        "</svg>",
    ])
    SVG_PATH.write_text("\n".join(parts), encoding="utf-8")


def write_result() -> None:
    depths, widths, overlaps = solve()

    workbook = load_workbook(WORKBOOK_PATH)
    sheet = workbook.active

    for col, value in enumerate(depths, start=2):
        cell = sheet.cell(row=2, column=col)
        cell.value = value
        cell.number_format = "0.00"

    for col, value in enumerate(widths, start=2):
        cell = sheet.cell(row=3, column=col)
        cell.value = value
        cell.number_format = "0.00"

    for col, value in enumerate(overlaps, start=2):
        cell = sheet.cell(row=4, column=col)
        cell.value = value
        if isinstance(value, float):
            cell.number_format = "0.00"

    workbook.save(WORKBOOK_PATH)
    draw_png(depths, widths, overlaps)
    draw_svg(depths, widths, overlaps)

    print("workbook:", WORKBOOK_PATH)
    print("png:", PNG_PATH)
    print("svg:", SVG_PATH)
    print("depths:", depths)
    print("widths:", widths)
    print("overlap_rates_percent:", overlaps)


if __name__ == "__main__":
    write_result()
